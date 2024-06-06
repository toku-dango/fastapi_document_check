from fastapi import FastAPI, File, UploadFile, Form, Request, HTTPException, Depends, Response
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from onelogin.saml2.auth import OneLogin_Saml2_Auth
from onelogin.saml2.utils import OneLogin_Saml2_Utils
from typing import List
import fitz
import openpyxl
import io
import docx


app = FastAPI()

# 静的ファイルのサービングのための設定
#app.mount("/templates", StaticFiles(directory="templates"), name="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# @app.get("/")
# def read_root():
#     return FileResponse('templates/index.html')

def get_current_user(request: Request):
    # RequestはFastAPIから提供されるクラスで、HTTPリクエストに関連するすべての情報を含んでいます。
    #このオブジェクトは、リクエストされたURL、HTTPメソッド（GET、POSTなど）、ヘッダー、パラメータ、ボディ、クッキーなどをアクセスするための属性やメソッドを持っています。
    user = request.session.get("user") #userキーの関連情報をsession属性を用いて取得する。sessionはユーザー固有のデータをサーバー側で保持するための辞書のようなストレージ
    if not user:
        return None
    return user

def init_saml_auth(req):
    auth = OneLogin_Saml2_Auth(req, custom_base_path="path_to_saml")
    return auth

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request, user=Depends(get_current_user)):
    #、Depends(get_current_user) は get_current_user 関数を実行して、その結果を user 変数に渡します。
    if not user:
        # ユーザーが認証されていない場合はログインページへリダイレクト
        return RedirectResponse(url="/api/saml/login")
    # 認証されている場合はindex.htmlを表示
    context = {"request": request, "user": user}
    return templates.TemplateResponse("index.html", context)

def prepare_saml_request(request: Request):
    url_data = {
        "https": "on" if request.url.scheme == "https" else "off",
        "http_host": request.url.netloc,
        "server_port": request.url.port,
        "script_name": request.url.path,
        "get_data": dict(request.query_params),
        "post_data": dict(request._form)
    }
    return url_data

@app.get("/api/saml/login")
def saml_login(request: Request):
    # リクエストデータを整形(SAML認証プロバイダ（OneLogin_Saml2_Auth）が要求する形式の辞書を作成します。)
    req = prepare_saml_request(request)#req はHTTPリクエストに関連するデータが辞書形式で渡されます。このデータには、リクエストされたURL、HTTPメソッド、クエリパラメータ、ヘッダー、POSTデータなどが含まれます。
    # SAML認証オブジェクトを初期化
    auth = init_saml_auth(req)# OneLogin_Saml2_Authは、OneLoginのPython SAMLライブラリの主要なクラスの一つで、SAML認証リクエストとレスポンスを処理します。custom_base_pathはSAML設定ファイル（たとえば、メタデータや証明書を含むファイル）が保存されているディレクトリのパスを指定します。
    # ログインURLにリダイレクト(IdP（Identity Provider）へのログインリダイレクトURLを生成します。このURLにユーザーをリダイレクトすることで、ユーザーはIdPのログインページで認証を行います。)
    return {"url": auth.login()}

@app.post("/api/saml/acs")
async def saml_acs(request: Request):
    #リクエストデータを整形します。
    req = prepare_saml_request(request)
    # 認証オブジェクトを初期化します。この時点で、IdPからのレスポンスを処理する準備ができます。
    auth = init_saml_auth(req)
    #IdPからのレスポンスを処理します。このメソッドはSAMLレスポンスの内容を解析し、ユーザーの認証状態を確認します。
    auth.process_response()
    #エラーをチェックし、エラーがなければユーザーをアプリケーションのメインページ（/static/index.html）にリダイレクトします。
    errors = auth.get_errors()
    if errors:
        raise HTTPException(status_code=401, detail="SAML authentication failed: " + ', '.join(errors))
    
    # 認証が成功したら、index.htmlにリダイレクト
    response = RedirectResponse(url="/static/index.html", status_code=303)
    return response

#File Import
@app.post("/upload-file")
async def create_upload_file(file: UploadFile = File(...)):
    print(file.content_type)
    try:
        if file.content_type == 'application/pdf':
            text = extract_text_from_pdf(await file.read())
            return {"filename": file.filename, "content": text}
        elif 'excel' in file.content_type or 'spreadsheet' in file.content_type:
            text = extract_data_from_excel(await file.read())
            return {"filename": file.filename, "content": text}
        elif file.content_type == 'text/plain':
            text = extract_text_from_txt(await file.read())
            return {"filename": file.filename, "content": text}
        elif 'document' in file.content_type or 'application/msword' in file.content_type:
            text = extract_text_from_docx(await file.read())
            return {"filename": file.filename, "content": text}
        else:
            return {"error": "Unsupported file type"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"message": str(e)})

def extract_text_from_pdf(contents):
    with fitz.open(stream=contents, filetype="pdf") as doc:
        text_output = ""
        for page_num, page in enumerate(doc, start=1):
            text_output += f"DocumentPage {page_num}\n"
            text_output += page.get_text() + "\n"
        return text_output

def extract_data_from_excel(contents):
   try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        text_output = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            text_output += f"SheetName: {sheet}\n"
            for row in worksheet.iter_rows(values_only=True):
                filtered_row = [cell for cell in row if cell is not None]
                if filtered_row:  # None だけの行を除外
                    text_output += "\t".join(map(str, filtered_row)) + "\n"
            text_output += "\n"
        return text_output
   except Exception as e:
        raise JSONResponse(status_code=500, content={"message": str(e)})

def extract_text_from_txt(contents):
    return contents.decode("utf-8")

def extract_text_from_docx(contents):
    doc = docx.Document(io.BytesIO(contents))
    text_output = ""
    for para in doc.paragraphs:
        text_output += para.text + "\n"
    return text_output

#Execute 
@app.post("/execute")
async def execute(request: Request):
    data = await request.json()
    input_text = data.get("inputText", "")
    additional_text = data.get("additionalText", "")
    check_points = data.get("checkPoints", "")
    result = f"入力テキスト: {input_text}\n追加のチェック観点: {additional_text}\nチェックポイント: {check_points}"
    print(result)
    return {"result": result}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
