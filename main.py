from fastapi import FastAPI, File, UploadFile, Form, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from typing import List
import fitz
import openpyxl
import io
import docx

app = FastAPI()

# 静的ファイルのサービングのための設定
#app.mount("/templates", StaticFiles(directory="templates"), name="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# @app.get("/")
# def read_root():
#     return FileResponse('templates/index.html')

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    context = {"request": request}
    return templates.TemplateResponse("index.html", context)

#File Import
@app.post("/upload-file")
async def create_upload_file(file: UploadFile = File(...)):
    print(file.content_type)
    try:
        if file.content_type == 'application/pdf':
            text = extract_text_from_pdf(await file.read())
            return {"filename": file.filename, "content": text}
        elif 'excel' in file.content_type or 'spreadsheet' in file.content_type:
            text = extract_data_from_excel(await file.read())
            return {"filename": file.filename, "content": text}
        elif file.content_type == 'text/plain':
            text = extract_text_from_txt(await file.read())
            return {"filename": file.filename, "content": text}
        elif 'document' in file.content_type or 'application/msword' in file.content_type:
            text = extract_text_from_docx(await file.read())
            return {"filename": file.filename, "content": text}
        else:
            return {"error": "Unsupported file type"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"message": str(e)})

def extract_text_from_pdf(contents):
    with fitz.open(stream=contents, filetype="pdf") as doc:
        text_output = ""
        for page_num, page in enumerate(doc, start=1):
            text_output += f"DocumentPage {page_num}\n"
            text_output += page.get_text() + "\n"
        return text_output

def extract_data_from_excel(contents):
   try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        text_output = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            text_output += f"SheetName: {sheet}\n"
            for row in worksheet.iter_rows(values_only=True):
                filtered_row = [cell for cell in row if cell is not None]
                if filtered_row:  # None だけの行を除外
                    text_output += "\t".join(map(str, filtered_row)) + "\n"
            text_output += "\n"
        return text_output
   except Exception as e:
        raise JSONResponse(status_code=500, content={"message": str(e)})

def extract_text_from_txt(contents):
    return contents.decode("utf-8")

def extract_text_from_docx(contents):
    doc = docx.Document(io.BytesIO(contents))
    text_output = ""
    for para in doc.paragraphs:
        text_output += para.text + "\n"
    return text_output

#Execute 
@app.post("/execute")
async def execute(request: Request):
    data = await request.json()
    input_text = data.get("inputText", "")
    additional_text = data.get("additionalText", "")
    check_points = data.get("checkPoints", "")
    result = f"入力テキスト: {input_text}\n追加のチェック観点: {additional_text}\nチェックポイント: {check_points}"
    print(result)
    return {"result": result}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
